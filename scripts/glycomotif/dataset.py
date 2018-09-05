
import tempfile,os.path,sys,re,itertools,traceback,copy
import gzip
import bz2
import warnings
import zipfile, shutil
from collections import defaultdict
try:
    import cPickle as Pickle
except ImportError:
    import Pickle

class Dataset(object):
    def __init__(self):
        self.tables = []
        self.namemap = {}
        self.current = None
    def new_table(self,name,table,rows):
        self.current = len(self.tables)
        self.tables.append(table)
        self.namemap[name] = self.current
        table.from_rows(rows)

class FileDataset(Dataset):
    def __init__(self,filename=None,names=None,tables=None):
        if not filename:
            filedesc,self.filename = tempfile.mkstemp()
            os.close(filedesc)
            self.tempfile = True
            assert names and tables
	    self.names_ = names
            self.from_tables(tables)
        else:
            self.filename = filename
            self.tempfile = False
            if not names and not tables:
                self.set_names()
            else:
	        self.names_ = names
		self.from_tables(tables)
        
import xlrd
import xlwt

class XLSFileDataset(FileDataset):
    def __init__(self,*args,**kw):
        self.colwidth=defaultdict(lambda :{})
        if kw.has_key('columnwidth'):
            self.colwidth.update(kw['columnwidth'])
            del kw['columnwidth']
        self.rowheight=defaultdict(lambda :{})
        if kw.has_key('rowheight'):
            self.rowheight.update(kw['rowheight'])
            del kw['rowheight']
        super(XLSFileDataset,self).__init__(*args,**kw)    
    underline = xlwt.easyxf('font: underline single')
    heightstyles={}
    @staticmethod
    def setcolwidth(ws,j,pixels):
        winpts = int(pixels*256.0/7.0);
        if ws.col(j).width < winpts:
            ws.col(j).width = winpts
    @staticmethod
    def setrowheight(ws,i,pixels):
        height = XLSFileDataset.heightstyles.get(pixels)
        if height == None:
            hinpts = int(pixels*12.0)
            height = xlwt.easyxf('font:height %d;'%hinpts)
            XLSFileDataset.heightstyles[pixels] = height
        ws.row(i).set_style(height)
    @staticmethod
    def writevalue(ws,i,j,value):
        if value == None:
            return
        try:
            value = Pickle.loads(value)
        except (AttributeError,Pickle.UnpicklingError,ValueError,EOFError,TypeError):
            pass
        if isinstance(value,dict):
            if value['type'] == 'url':
                ws.write(i,j,
                         xlwt.Formula('HYPERLINK("%s";"%s")'%
                                      (value['url'],value['text'])),
                         XLSFileDataset.underline)
                return
            elif value['type'] == 'image':
                ws.insert_bitmap(value['image'],i,j)
                return
        try:
            ws.write(i,j,float(value))
            return
        except (ValueError,TypeError):
            pass
        try:
            ws.write(i,j,int(value))
            return
        except (ValueError,TypeError):
            pass
        ws.write(i,j,str(value))
    def set_names(self):
	book = xlrd.open_workbook(self.filename)
        for sheet_idx,sheet_name in enumerate(book.sheet_names()):
	    self.names_.append(sheet_name)
    def from_tables(self,tables):
        wb = xlwt.Workbook()
        for n,t in zip(self.names_,tables):
	    ws = wb.add_sheet(n)
            dcw = self.colwidth[n].get(None)
            for j,h in enumerate(t.headers()):
                cw = self.colwidth[n].get(h,dcw)
                if cw != None:
                    XLSFileDataset.setcolwidth(ws,j,cw)
            for j,h in enumerate(t.headers()):
		XLSFileDataset.writevalue(ws,0,j,h)
            drh = self.rowheight[n].get(None)
            for i,r in enumerate(t.rows()):
	        if i >= 65535:
		    break
                rh = self.rowheight[n].get(i,drh)
                if rh:
                    XLSFileDataset.setrowheight(ws,i+1,rh)
                for j,h in enumerate(t.headers()):
                    XLSFileDataset.writevalue(ws,i+1,j,r[h])
        wb.save(self.filename)

import openpyxl

class XLSXFileDataset(FileDataset):
    def __init__(self,*args,**kw):
        self.colwidth=defaultdict(lambda :{})
        if kw.has_key('columnwidth'):
            self.colwidth.update(kw['columnwidth'])
            del kw['columnwidth']
        self.rowheight=defaultdict(lambda :{})
        if kw.has_key('rowheight'):
            self.rowheight.update(kw['rowheight'])
            del kw['rowheight']
	self._highlight=None
        self._highlight1=None
        if kw.has_key('highlight'):
            self._highlight = openpyxl.style.Color(kw['highlight'])
            del kw['highlight']
            self._highlight1 = self._highlight1
        if kw.has_key('highlight1'):
            self._highlight1 = openpyxl.style.Color(kw['highlight1'])
            del kw['highlight1']
        super(XLSXFileDataset,self).__init__(*args,**kw)    
    @staticmethod
    def setcolwidth(ws,j,pixels):
        winpts = pixels*0.14214 # openpyxl.shared.units.pixels_to_points(pixels)
        ws.cell(row=0,column=j)
        ws.column_dimensions[openpyxl.cell.get_column_letter(j+1)].width = winpts
    @staticmethod
    def setrowheight(ws,i,pixels):
        hinpts = pixels*3./4. # openpyxl.shared.units.pixels_to_points(pixels);
        ws.cell(row=i,column=0)
        ws.row_dimensions[i+1].height = hinpts

    @staticmethod
    def highlight(ws,i,j,c=None):
        if c != None:
            ws.cell(row=i,column=j).style.fill.fill_type = openpyxl.style.Fill.FILL_SOLID
            ws.cell(row=i,column=j).style.fill.start_color = c
            ws.cell(row=i,column=j).style.fill.end_color = c

    @staticmethod
    def header(ws,j,c=None):
        ws.cell(row=0,column=j).style.font.bold = True
        ws.cell(row=0,column=j).style.font.size = 14
        ws.cell(row=0,column=j).style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
        XLSXFileDataset.highlight(ws,0,j,c)
        
    @staticmethod
    def writevalue(ws,i,j,value):
        if value == None:
            return
        try:
            value = Pickle.loads(value)
        except (Pickle.UnpicklingError,ValueError,EOFError,TypeError,AttributeError):
            pass
        if isinstance(value,dict):
            if value['type'] == 'url':
                ws.cell(row=i,column=j).value = value['text']
                ws.cell(row=i,column=j).hyperlink = value['url']
                ws.cell(row=i,column=j).style.font.underline = openpyxl.style.Font.UNDERLINE_SINGLE
                return
            if value['type'] == 'image':
                img = openpyxl.drawing.Image(value['image'])
                img.anchor(ws.cell(row=i,column=j),type=openpyxl.drawing.Image.ONECELL)
                ws.add_image(img)
                return
        try:
            ws.cell(row=i,column=j).value=float(value)
            return
        except (ValueError,TypeError):
            pass
        try:
            ws.cell(row=i,column=j).value=int(value)
            return
        except (ValueError,TypeError):
            pass
        ws.cell(row=i,column=j).value=str(value)        
    
    # underline = xlwt.easyxf('font: underline single')
    def set_names(self):
        book = openpyxl.reader.excel.load_workbook(filename = self.filename)
        for sheet_idx,sheet_name in enumerate(book.get_sheet_names()):
            self.names_.append(sheet_name)

    def from_tables(self,tables):
        wb = openpyxl.workbook.Workbook()
        for i,(n,t) in enumerate(zip(self.names_,tables)):
            if i == 0:
                ws = wb.worksheets[0]
            else:
                ws = wb.create_sheet()
            ws.title = n
            dcw = self.colwidth[n].get(None)
            for j,h in enumerate(t.headers()):
                cw = self.colwidth[n].get(h,dcw)
                if cw != None:
                    XLSXFileDataset.setcolwidth(ws,j,cw)
	    for j,h in enumerate(filter(lambda h: not h.startswith('_'),t.headers())):
		ws.cell(row=0,column=j).value=h
                XLSXFileDataset.header(ws,j,self._highlight)
            drh = self.rowheight[n].get(None)
            for i,r in enumerate(t.rows()):
                rh = self.rowheight[n].get(i,drh)
                if rh:
                    XLSXFileDataset.setrowheight(ws,i+1,rh)
                for j,h in enumerate(filter(lambda h: not h.startswith('_'),t.headers())):
                    XLSXFileDataset.writevalue(ws,i+1,j,r.get(h))
                    ws.cell(row=i+1,column=j).style.font.size = 12
                    if r.get('_highlight',False):
                        XLSXFileDataset.highlight(ws,i+1,j,self._highlight1)
                        
        wb.save(self.filename)

class ZIPFileDataset(FileDataset):

    def from_tables(self,tables):
        tmpdir = tempfile.mkdtemp()
	for n,t in zip(self.names_,tables):
            to = CSVFileTable(filename=os.path.join(tmpdir,n+'.csv'),
                              from_rows=t,compression=None)
        zf = zipfile.ZipFile(self.filename,'w',compression=zipfile.ZIP_DEFLATED,allowZip64=True)
        for n in self.names_:
            zf.write(os.path.join(tmpdir,n+'.csv'),n+'.csv')
        zf.close()
        shutil.rmtree(tmpdir,ignore_errors=True)
        
class Table(object):

    def extract(self,*keys):
        for r in self.rows():
            if len(keys) == 1:
                yield r[keys[0]]
            else:
                yield [ r[k] for k in keys ]

    def headers(self):
        return self.headers_

    def __iter__(self):
        return self.rows()

class MemoryTable(Table):

    def __init__(self,headers):
        self.headers_ = list(headers)
        self.therows = []

    def rows(self):
        for r in self.therows:
            yield dict(zip(self.headers_,r))

    def size(self):
	return len(self.therows)

    def from_rows(self,rows):
        self.therows = []
        for r in rows:
            self.therows.append([ r.get(h,"") for h in self.headers_ ])

    def sort(self,key=None,cmp=None):
        if key:
            self.therows.sort(key=lambda r: key(dict(zip(self.headers_,r))))
        else:
            self.therows.sort(cmp=lambda a,b: cmp(dict(zip(self.headers_,a)),dict(zip(self.headers_,b))))

class FileTable(Table):

    def __init__(self,filename=None,headers=None,from_rows=None,compression='bz2'):
	self.compression = compression
        if not filename:
            cmp = ''
            if compression:
                cmp = '.'+compression
            filedesc,self.filename = tempfile.mkstemp(suffix=cmp)
            os.close(filedesc)
            self.tempfile = True
            if from_rows:
                self.headers_ = from_rows.headers()
                self.from_rows(from_rows)
            else:
                assert(headers)
                self.headers_ = headers
        else:
            self.filename = filename
            self.tempfile = False
            if not headers and not from_rows:
                self.set_headers()
            else:
                if from_rows:
                    self.headers_ = from_rows.headers()
                    self.from_rows(from_rows)
                else:
                    self.headers_ = headers
        
    def __del__(self):
        if hasattr(self,'tempfile') and self.tempfile:
            if os.path.exists(self.filename):
                os.unlink(self.filename)

    def open(self,mode='r'):
	if not isinstance(self.filename,basestring):
	    return iter(self.filename)
        if self.filename.lower().endswith(".gz") or \
               self.compression == "gz" or \
               os.path.exists(self.filename+".gz"):
            try:
                if 'r' in mode:
                    if self.filename.lower().endswith(".gz"):
                        h = gzip.open(self.filename,mode)
                        h.readline()
                        h.seek(0)
                    else:
                        h = gzip.open(self.filename+".gz",mode)
                        h.readline()
                        h.seek(0)
                        self.filename += ".gz"
                else:
                    if self.filename.lower().endswith(".gz"):
                        h = gzip.open(self.filename,mode)
                    else:
                        h = gzip.open(self.filename+".gz",mode)
                        self.filename += ".gz"
                return h
            except IOError:
                return open(self.filename,mode)
	elif self.filename.lower().endswith(".bz2") or \
                 self.compression == "bz2" or \
                 os.path.exists(self.filename+".bz2"):
            try:
                if 'r' in mode:
                    if self.filename.lower().endswith(".bz2"):
                        h = bz2.BZ2File(self.filename,mode)
                        h.readline()
                        h.seek(0)
                    else:
                        h = bz2.BZ2File(self.filename+".bz2",mode)                        
                        h.readline()
                        h.seek(0)
                        self.filename += ".bz2"
                else:
                    if self.filename.lower().endswith(".bz2"):
                        h = bz2.BZ2File(self.filename,mode)
                    else:
                        h = bz2.BZ2File(self.filename+".bz2",mode)
                        self.filename += ".bz2"
                return h
            except IOError, e:
                # print >>sys.stderr, e
                return open(self.filename,mode+'U')
	else:
	    return open(self.filename,mode+'U')

import csv

class CSVFileTable(FileTable):
    dialect = csv.excel

    def set_headers(self):
        h = self.open()
        t = csv.reader(h,dialect=self.dialect)
        self.headers_ = t.next()
        h.close()
    
    def rows(self):
        h = self.open()
        rows = csv.DictReader(h,dialect=self.dialect)
        for r in rows:
            yield r
        h.close()

    def from_rows(self,rows):
        wh = self.open(mode='wb')
        dwh = csv.DictWriter(wh,self.headers_,extrasaction='ignore',dialect=self.dialect)
        dwh.writerow(dict([(h,h) for h in self.headers_]))
        for r in rows:
            dwh.writerow(r)
        wh.close()

class TSVFileTable(CSVFileTable):
    dialect = csv.excel_tab

import xlrd
import xlwt

class XLSFileTable(FileTable):
    def __init__(self,*args,**kw):
        if kw.has_key('sheet'):
            if kw['sheet']:
                self.sheet = kw['sheet']
            else:
                self.sheet = None
            del kw['sheet']
        else:
            self.sheet = None
        self.colwidth={}
        if kw.has_key('columnwidth'):
            self.colwidth = kw['columnwidth']
            del kw['columnwidth']
        self.rowwidth={}
        if kw.has_key('rowheight'):
            self.rowheight = kw['rowheight']
            del kw['rowheight']
        super(XLSFileTable,self).__init__(*args,**kw)
        
    def set_headers(self):
        self.headers_ = []
        book = xlrd.open_workbook(self.filename)
        for sheet_idx,sheet_name in enumerate(book.sheet_names()):
            if self.sheet and sheet_name != self.sheet:
                continue
            if not self.sheet and sheet_idx != 0:
                continue
            sheet = book.sheet_by_name(sheet_name)
            for v in sheet.row_values(0):
                if v in (None,""):
                    v = None
                self.headers_.append(v)
	assert len(self.headers_) > 0

    def rows(self):
        book = xlrd.open_workbook(self.filename,formatting_info=True)
        # print repr(book.xf_list)
        for sheet_idx,sheet_name in enumerate(book.sheet_names()):
            if self.sheet != None and sheet_name != self.sheet:
                continue
            if self.sheet == None and sheet_idx != 0:
                continue
            sheet = book.sheet_by_name(sheet_name)
            for r in range(sheet.nrows):
                if r == 0:
                    continue
                yield dict((h,self.cell2value(c)) for (h,c) in zip(self.headers_,sheet.row_slice(r)) if h != None)

    def cell2value(self,c):
        # print repr(c),repr(c.ctype),repr(c.value),repr(c.xf_index)
        return c.value
        
    def from_rows(self,rows):
        wb = xlwt.Workbook()
        assert(self.sheet != None)
        ws = wb.add_sheet(self.sheet)
        dcw = self.colwidth.get(None)
        for j,h in enumerate(self.headers_):
            cw = self.colwidth.get(h,dcw)
            if cw != None:
                XLSFileDataset.setcolwidth(ws,j,cw)
        for j,h in enumerate(self.headers_):
            XLSFileDataset.writevalue(ws,0,j,h)
        drh = self.rowheight.get(None)
        for i,r in enumerate(rows):
            rh = self.rowheight.get(i,drh)
            if rh:
                XLSFileDataset.setrowheight(ws,i+1,rh)
            for j,h in enumerate(self.headers_):
                XLSFileDataset.writevalue(ws,i+1,j,r[h])
        wb.save(self.filename)

import openpyxl

class XLSXFileTable(FileTable):
    def __init__(self,*args,**kw):
        if kw.has_key('sheet'):
            if kw['sheet']:
                self.sheet = kw['sheet']
            else:
                self.sheet = None
            del kw['sheet']
        else:
            self.sheet = None
        super(XLSXFileTable,self).__init__(*args,**kw)
        
    def set_headers(self):
        self.headers_ = []
        book = openpyxl.reader.excel.load_workbook(filename = self.filename)
        for sheet_idx,sheet_name in enumerate(book.get_sheet_names()):
            if self.sheet and sheet_name != self.sheet:
                continue
            if not self.sheet and sheet_idx != 0:
                continue
            sheet = book.get_sheet_by_name(sheet_name)
            ncol =  sheet.get_highest_column()
            for i in range(ncol):
                h = sheet.cell(row=0,column=i).value
                if h in (None,""):
                    h = None
                self.headers_.append(h)
	assert len(self.headers_) > 0

    def rows(self):
        book = openpyxl.reader.excel.load_workbook(filename = self.filename)
        for sheet_idx,sheet_name in enumerate(book.get_sheet_names()):
            if self.sheet != None and sheet_name != self.sheet:
                continue
            if self.sheet == None and sheet_idx != 0:
                continue
            sheet = book.get_sheet_by_name(sheet_name)
            ncol = sheet.get_highest_column()
            nrow = sheet.get_highest_row()
            for r in range(nrow):
                if r == 0:
                    continue
                row = []
                for c in range(ncol):
                    value = sheet.cell(row=r,column=c).value
                    row.append(value)
                yield dict((h,v) for (h,v) in zip(self.headers_,row) if h != None)
        
    def from_rows(self,rows):
        wb = openpyxl.workbook.Workbook()
        assert(self.sheet != None)
        ws = wb.worksheets[0]
        ws.title = self.sheet
        for j,h in enumerate(self.headers_):
            ws.cell(row=0,column=j).value=h
        for i,r in enumerate(rows):
            for j,h in enumerate(self.headers_):
                XLSXFileDataset.writevalue(ws,i+1,j,r.get(h))		
        wb.save(filename=self.filename)

# Reader only...
class MassLynxFileTable(FileTable):

    def set_headers(self):
        self.headers_ = filter(None,"""
            rt
            intensity
            area
            start_rt
            end_rt
            dummy1
            dummy2
            GU
        """.split())
    
    def rows(self):
        h = self.open()
        for r in h:
            yield dict(zip(self.headers_,r.split()))
        h.close()    

class AddField(Table):

    def __init__(self,tablein,fieldname,valuefunction):
        self.tin = tablein
        self.fieldname = fieldname
        self.headers_ = list(tablein.headers())
        self.headers_.append(fieldname)
        self.f = valuefunction

    def rows(self):
        for r in self.tin:
            r[self.fieldname] = self.f(r)
            yield r

class SetField(Table):

    def __init__(self,tablein,fieldname,valuefunction):
        self.tin = tablein
        self.headers_ = list(tablein.headers())
        self.fieldname = fieldname
        self.f = valuefunction

    def rows(self):
        for r in self.tin:
            r[self.fieldname] = self.f(r)
            yield r

class AddFields(Table):
    def __init__(self,tablein,fieldfunction,valuefunction):
        self.tin = tablein
        self.headers_ = fieldfunction(tablein.headers())
        self.f = valuefunction

    def rows(self):
        for r in self.tin:
            yield self.f(r)

class ReplaceFromList(Table):
    def __init__(self,tablein,field,values):
        self.tin = tablein
        self.headers_ = tablein.headers()
        assert field in self.headers_
        self.f = field
        self.l = values

    def rows(self):
        for r,v in zip(self.tin,self.l):
            r[self.f] = v
            yield r

class FilterRowsByFunction(Table):

    def __init__(self,tablein,fn):
        self.tin = tablein
        self.headers_ = tablein.headers()
        self.f = fn

    def rows(self):
        for r in self.tin:
            if self.f(r):
                yield r

class FilterRows(Table):

    def __init__(self,tablein,expr):
        self.tin = tablein
        self.headers_ = tablein.headers()
        self.expr = expr
        self.varmap = {}
        self.listmap = {}
        j = 0;
        for h in sorted(self.headers_,key=lambda h: -len(h)):
            if h in self.expr:
                var = "v%d"%j
                self.expr = self.expr.replace(h,var)
                self.varmap[h] = var
                j += 1
        for h in sorted(self.headers_,key=lambda h: -len(h)):
            m = re.search(r'^(.*?)\d+$',h)
            if m and m.group(1) in self.expr:
                var = "v%d"%j
                self.expr = self.expr.replace(m.group(1),var)
                self.listmap[m.group(1)] = var
                j += 1

    def rows(self):
        for r in self.tin:
            locals = {}
            for h in self.listmap:
                locals[self.listmap[h]] = []
                any = False
                for i in range(21):
                    try:
                        locals[self.listmap[h]].append(float(r["%s%d"%(h,i)]))
                        any = True
                    except KeyError:
                        if not any:
                            locals[self.listmap[h]].append(None)
                        else:
                            break
            for h in self.varmap:
                try:
                    locals[self.varmap[h]] = float(r[h])
                except ValueError:
                    locals[self.varmap[h]] = r[h]
            # print >>sys.stderr, self.expr,locals
            if eval(self.expr,globals(),locals):
                yield r

class HeaderMap(Table):
    def __init__(self,tablein,headermap):
        self.tin = tablein
        self.headers_ = map(headermap,tablein.headers())
        self.headermap = headermap

    def rows(self):
        for r in self.tin:
            yield dict( [ (self.headermap(k),v) for (k,v) in r.items() ] )
            
class ColumnSelect(Table):
    def __init__(self,tablein,goodset):
        self.tin = tablein
        self.headers_ = [ h for h in tablein.headers() if h in goodset ]
        self.goodset = goodset

    def rows(self):
        for r in self.tin:
            yield dict( [ (k,v) for (k,v) in r.items() if k in self.goodset ] )
            
class ColumnRemove(Table):
    def __init__(self,tablein,badset):
        self.tin = tablein
        self.headers_ = [ h for h in tablein.headers() if h not in badset ]
        self.badset = badset

    def rows(self):
        for r in self.tin:
            yield dict( [ (k,v) for (k,v) in r.items() if k not in self.badset ] )
            
class ColumnRemoveRegex(Table):
    def __init__(self,tablein,badregex):
        self.tin = tablein
        self.headers_ = [ h for h in tablein.headers() if not re.search(badregex,h) ]

    def rows(self):
        for r in self.tin:
            yield dict( [ (k,v) for (k,v) in r.items() if k in self.headers_ ] )

class ValueMap(Table):
    def __init__(self,tablein,mapfn,colset=None):
        self.tin = tablein
        self.headers_ = list(tablein.headers())
        self.f = mapfn
	self.colset = colset

    def applymap(self,item):
	if self.colset == None or item[0] in self.colset:
	    return item[0],self.f(item[1])
	return item

    def rows(self):
        for r in self.tin:
            yield dict(map(self.applymap,r.items()))

class MoveField(Table):
    def __init__(self,tablein,pos,cols,after=True):
	self.tin = tablein
	self.headers_ = list(tablein.headers())
	for c in cols:
	    assert c in self.headers_
	try:
	    p = int(pos)
 	except (TypeError, ValueError):
	    assert pos in self.headers_
	    p = self.headers_.index(pos)
	for c in reversed(cols):
	    p0 = self.headers_.index(c)
	    del self.headers_[p0]
	    if after:
	        self.headers_.insert(p+1,c)
	    else:
	        self.headers_.insert(p,c)

    def rows(self):
        for r in self.tin:
            yield r

class ParsimonyCSV:
    def __init__(self,proteins):
	self._proteins = set(proteins)
    def rows(self,inrows):
	for r in inrows:
	    proteins = set(r['protein'].split(';'))
	    proteins &= self._proteins
	    if len(proteins) > 0:
		r['protein'] = ';'.join(proteins)
		yield r
    def rewrite(self,infile,outfile):
	rows = CSVFileTable(filename=infile)
	newrows = CSVFileTable(filename=outfile,headers=rows.headers(),compression=None)
	newrows.from_rows(self.rows(rows))

